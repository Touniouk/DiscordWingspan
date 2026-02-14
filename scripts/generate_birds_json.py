#!/usr/bin/env python3
"""
Generate birds.json from the canonical wingspan xlsx spreadsheet.

Usage:
    python3 scripts/generate_birds_json.py

Requires openpyxl:
    pip install openpyxl
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl is required. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
XLSX_PATH = PROJECT_ROOT / "src" / "main" / "resources" / "wingspan-20260128.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "src" / "main" / "resources" / "birds.json"

# xlsx "Set" values -> JSON "Expansion" values
# Only "core" needs remapping; all others (european, oceania, asia,
# americas, promoAsia, promoCA, promoEurope, promoNZ, promoUK, promoUS)
# pass through as-is.
EXPANSION_MAP = {
    "core": "originalcore",
}

# Columns that should be renamed from xlsx -> JSON
RENAME = {
    "Set": "Expansion",
    "Egg limit": "Egg capacity",
}

# Columns whose string values should be title-cased
TITLE_CASE_COLS = {"Color", "Nest type"}

# Numeric columns (value is float or null)
NUMERIC_COLS = {
    "Victory points", "Egg capacity", "Total food cost",
    "Invertebrate", "Seed", "Fish", "Fruit", "Rodent", "Nectar", "Wild (food)",
}

# Flag columns ("X" or null)
FLAG_COLS = {
    "Predator", "Flocking", "Bonus card",
    "Forest", "Grassland", "Wetland",
    "/ (food cost)", "* (food cost)",
    "Swift Start", "Automa ban", "Fan Art Pack?",
    # Region columns
    "North America", "Central America", "South America",
    "Europe", "Asia", "Africa", "Oceania",
    # Bonus card columns
    "Anatomist", "Cartographer", "Historian", "Photographer",
    "Backyard Birder", "Bird Bander", "Bird Counter", "Bird Feeder",
    "Diet Specialist", "Enclosure Builder", "Endangered Species Protector",
    "Falconer", "Fishery Manager", "Food Web Expert", "Forester",
    "Large Bird Specialist", "Nest Box Builder", "Omnivore Expert",
    "Passerine Specialist", "Platform Builder", "Prairie Manager",
    "Rodentologist", "Small Clutch Specialist", "Viticulturalist",
    "Wetland Scientist", "Wildlife Gardener",
}

# String columns (value or null)
STRING_COLS = {
    "Common name", "Scientific name", "Expansion", "Color",
    "Power text", "Flavor text", "Nest type", "Wingspan",
    "Beak direction", "Fan Art flavor text", "Fan art beak direction",
}

# Ordered list of JSON output keys (matching old birds.json field order,
# plus new fields from xlsx that weren't in the old JSON)
OUTPUT_KEY_ORDER = [
    "Common name",
    "Scientific name",
    "Expansion",
    "Color",
    "Power text",
    "Flavor text",
    "Predator",
    "Flocking",
    "Bonus card",
    "Victory points",
    "Nest type",
    "Egg capacity",
    "Wingspan",
    "Forest",
    "Grassland",
    "Wetland",
    "Invertebrate",
    "Seed",
    "Fish",
    "Fruit",
    "Rodent",
    "Nectar",
    "Wild (food)",
    "/ (food cost)",
    "* (food cost)",
    "Total food cost",
    "Beak direction",
    "Swift Start",
    "Automa ban",
    "North America",
    "Central America",
    "South America",
    "Europe",
    "Asia",
    "Africa",
    "Oceania",
    "Fan Art Pack?",
    "Fan Art flavor text",
    "Fan art beak direction",
    "Anatomist",
    "Cartographer",
    "Historian",
    "Photographer",
    "Backyard Birder",
    "Bird Bander",
    "Bird Counter",
    "Bird Feeder",
    "Diet Specialist",
    "Enclosure Builder",
    "Endangered Species Protector",
    "Falconer",
    "Fishery Manager",
    "Food Web Expert",
    "Forester",
    "Large Bird Specialist",
    "Nest Box Builder",
    "Omnivore Expert",
    "Passerine Specialist",
    "Platform Builder",
    "Prairie Manager",
    "Rodentologist",
    "Small Clutch Specialist",
    "Viticulturalist",
    "Wetland Scientist",
    "Wildlife Gardener",
]


def convert_value(json_key, raw_value):
    """Convert a raw xlsx cell value to the appropriate JSON value."""
    if json_key in NUMERIC_COLS:
        if raw_value is None or raw_value == "":
            return None
        if isinstance(raw_value, str):
            # Formula cells (e.g. Total food cost) - skip, handled separately
            return None
        return float(raw_value)

    if json_key in FLAG_COLS:
        if raw_value == "X":
            return "X"
        return None

    if json_key in STRING_COLS:
        if raw_value is None or raw_value == "":
            # Nest parasites (cuckoos, cowbirds) have no nest type;
            # the old JSON used the string "None" for these
            if json_key == "Nest type":
                return "None"
            return None
        val = str(raw_value)
        if json_key in TITLE_CASE_COLS:
            val = val.title()
        return val

    return raw_value


def compute_total_food_cost(row_data):
    """Compute total food cost from individual food columns.

    If '/ (food cost)' is 'X', the cost is 1 (any one of the listed foods).
    Otherwise, sum all individual food amounts.
    """
    if row_data.get("/ (food cost)") == "X":
        return 1.0

    food_cols = ["Invertebrate", "Seed", "Fish", "Fruit", "Rodent", "Nectar", "Wild (food)"]
    total = 0.0
    for col in food_cols:
        val = row_data.get(col)
        if val is not None:
            total += val
    return total


def compute_expansion(set_value, swift_start):
    """Map xlsx Set + Swift Start to JSON Expansion value."""
    if swift_start == "X" and set_value == "core":
        return "swiftstart"
    return EXPANSION_MAP.get(set_value, set_value)


def main():
    if not XLSX_PATH.exists():
        print(f"Error: xlsx file not found at {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["Birds"]

    headers = [cell.value for cell in ws[1]]
    birds = []

    for row in ws.iter_rows(min_row=2):
        raw = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                raw[headers[i]] = cell.value

        # Skip empty rows
        if not raw.get("Common name"):
            continue

        # Build the bird dict with proper JSON key names
        bird = {}
        for xlsx_col, value in raw.items():
            json_key = RENAME.get(xlsx_col, xlsx_col)
            bird[json_key] = convert_value(json_key, value)

        # Handle Expansion specially (Set + Swift Start)
        bird["Expansion"] = compute_expansion(
            raw.get("Set"), raw.get("Swift Start")
        )

        # Compute Total food cost (xlsx has formulas, data_only may return None)
        if bird.get("Total food cost") is None:
            bird["Total food cost"] = compute_total_food_cost(bird)

        # Wingspan should be a string (integer string, or "*" for flightless)
        if bird.get("Wingspan") is not None:
            ws_val = bird["Wingspan"]
            try:
                bird["Wingspan"] = str(int(float(ws_val)))
            except (ValueError, TypeError):
                bird["Wingspan"] = str(ws_val)

        # Build ordered output dict
        ordered = {}
        for key in OUTPUT_KEY_ORDER:
            if key in bird:
                ordered[key] = bird[key]

        birds.append(ordered)

    wb.close()

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(birds, f, indent=4, ensure_ascii=False)

    print(f"Generated {OUTPUT_PATH} with {len(birds)} birds")


if __name__ == "__main__":
    main()
